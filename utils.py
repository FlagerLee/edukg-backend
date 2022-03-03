from asyncio import exceptions
from pydantic import BaseModel, validator
from openpyxl import load_workbook
from lxml import etree
from PIL import Image
from paddleocr import PaddleOCR
import pybase64
import sys
import os
import re
import json
import time
import requests


class result(BaseModel):
    exam_id: str
    question_id: str
    result: str
    
    @validator('exam_id')
    def exam_id_re(cls, eid):
        res = re.match(r'^(\d+)_(\d+)$', eid)
        if not res:
            raise ValueError('exam id value error')
        from datetime import datetime
        eid_year = int(res.group(1))
        if eid_year < 2013 or eid_year > datetime.now().year:
            raise ValueError('exam year out of range')
        exam_path = '/'.join([json_root_path, str(eid_year) + 'GaoKao', eid])
        if not os.path.exists(exam_path):
            raise ValueError('exam id does not exist')
        return eid

    @validator('question_id')
    def question_id_re(cls, qid):
        res = re.match(r'^(\d+)_(\d+)$', qid)
        if not res:
            raise ValueError('exam id value error')
        return qid

    @validator('result')
    def result_re(cls, result):
        res = json.loads(result)
        assert type(res) is dict
        assert res.__contains__('ID')
        assert res.__contains__('Subject')
        assert res.__contains__('Grade')
        assert res.__contains__('School')
        assert res.__contains__('SourceLink')
        assert res.__contains__('TestName')
        assert res.__contains__('Content')
        assert res.__contains__('Questions')
        return res


# 图片内容，使用Base64加密和解密。mood只能为'sentence'/'latex'，分别使用普通ocr和图片转latex进行
class ocr_img(BaseModel):
    img_url: str
    mood: str

    @validator('mood')
    def mood_re(cls, mood):
        assert mood == 'sentence' or mood == 'latex'
        return mood

json_root_path = None
html_root_path = None
excel_root_path = None


def load_cfg():
    cfg_path = 'settings.cfg'
    with open(cfg_path, 'r') as f:
        for line in f.readlines():
            cfg_name, cfg_path = line.split('=')
            cfg_name = cfg_name.lower()
            if cfg_name in globals():
                globals()[cfg_name] = cfg_path.strip('\n')


def scan_files():
    assert 'excel_root_path' in globals()
    global excel_root_path
    assert excel_root_path is not None
    exam_list = []
    for excel_name in os.listdir(excel_root_path):
        if excel_name.endswith('.xlsx'):
            # is excel file
            work_sheet = load_workbook('/'.join([excel_root_path, excel_name])).active
            for row in work_sheet.iter_rows(min_row=2):
                exam_list.append({
                    'exam_name': row[0].value,
                    'exam_url': row[1].value,
                    'exam_year': row[2].value,
                    'exam_subject': row[3].value,
                    'exam_grade': row[4].value,
                    'exam_id': row[5].value
                })
    return exam_list


def list_questions(exam_id: str):
    year = exam_id.split('_')[0]
    htmls_path = '/'.join([html_root_path, year + 'GaoKao', exam_id])
    html_list = os.listdir(htmls_path)
    qid_list = []
    for html_name in html_list:
        if not html_name.endswith('.html'):
            continue
        qid = '_'.join(html_name.split('.')[0].split('_')[:2])
        if qid not in qid_list:
            qid_list.append(qid)
    #qid_list.sort(cmp=lambda str1, str2: int(str1.split('_')[-1]) < int(str2.split('_')[-1]))
    # 为什么cmp函数不起作用？
    def cmp(str1, str2):
        print(str1)
        print(str2)
        print(int(str1.split('_')[-1]) < int(str2.split('_')[-1]))
        return int(str1.split('_')[-1]) < int(str2.split('_')[-1])
    #return sorted(qid_list, key=cmp_to_key(cmp))
    return sorted(qid_list)


def get_single_question(exam_id: str, question_id: str):
    year = exam_id.split('_')[0]
    htmls_path = '/'.join([html_root_path, year + 'GaoKao', exam_id])
    question_html_path = '/'.join([htmls_path, question_id + '.html'])
    answer_html_path = '/'.join([htmls_path, question_id + '_answer.html'])
    result_json_path = '/'.join([json_root_path, year + 'GaoKao', exam_id, question_id + '.json'])
    # then find all pictures in question html and answer html
    img_list = []
    parser = etree.HTMLParser(encoding='utf-8')
    question_tree = etree.parse(question_html_path, parser=parser)
    result_src = question_tree.xpath('//img/@src')
    result_data_lazysrc = question_tree.xpath('//img/@data-lazysrc')
    img_list.extend(result_src)
    img_list.extend(result_data_lazysrc)
    answer_tree = etree.parse(answer_html_path, parser=parser)
    result_src = answer_tree.xpath('//img/@src')
    result_data_lazysrc = answer_tree.xpath('//img/@data-lazysrc')
    img_list.extend(result_src)
    img_list.extend(result_data_lazysrc)
    def read_file(path):
        if not os.path.exists(path):
            return None
        with open(path, 'r') as f:
            return ''.join(f.readlines())
    return {
        'problem_html': read_file(question_html_path),
        'answer_html': read_file(answer_html_path),
        'result': read_file(result_json_path),
        'pics': img_list
    }


class HiddenPrints:
    def __init__(self, activated=True):
        # activated参数表示当前修饰类是否被**
        self.activated = activated
        self.original_stdout = None

    def open(self):
        sys.stdout.close()
        sys.stdout = self.original_stdout

    def close(self):
        self.original_stdout = sys.stdout
        sys.stdout = open(os.devnull, 'w')

    def __enter__(self):
        if self.activated:
            self.close()

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.activated:
            self.open()


def download_img(url):
    file_name = url.split('/')[-1]
    res = requests.get(url)
    with open(file_name, 'wb') as f:
        f.write(res.content)
    img = remove_transparency(Image.open(file_name))
    if img.mode == 'P':
        img = img.convert('RGB')
    return img, file_name


def remove_transparency(img_pil, bg_color=(255, 255, 255)):
    if img_pil.mode in ('RGBA', 'LA'):
        alpha = img_pil.split()[-1]
        bg = Image.new('RGBA', img_pil.size, bg_color + (255,))
        bg.paste(img_pil, mask=alpha)
        return bg.convert('RGB')
    else:
        return img_pil


def convert(img_md: ocr_img): # md: meta data
    img, img_path = download_img(img_md.img_url)
    height, width = img.size
    if img_md.mood == 'sentence':
        if height <= 100 and width <= 100:
            # 将图像横向拼接识别文字
            new_img = Image.new(img.mode, (10 * (img.size[0] + 25), img.size[1] + 100))
            for i in range(10):
                new_img.paste(img, (12 + i * (img.size[0] + 25), 50))
            img.close()
            new_img.save(img_path)
            new_img.close()
            hide_print = HiddenPrints()
            hide_print.close()
            ocr = PaddleOCR(use_angle_cls=True, lang='ch')
            result = ocr.ocr(img_path, cls=True)
            hide_print.open()
            result_str = ''
            for line in result:
                if float(line[1][1]) > 0.8:
                    result_str += line[1][0]
            if len(result_str) % 10 != 0:
                return ''
            return result_str[:len(result_str) // 10]
        else:
            img.close()
            hide_print = HiddenPrints()
            hide_print.close()
            ocr = PaddleOCR(use_angle_cls=True, lang='ch')
            result = ocr.ocr(img_path, cls=True)
            hide_print.open()
            result_str = ''
            for line in result:
                result_str += line[1][0]
            return result_str
    else:
        img.close()
        with open(img_path, 'rb') as f:
            Image_data = pybase64.b64decode(f.read())
        url = "https://www.bing.com/cameraexp/api/v1/getlatex"
        headers = {
            "Host": "www.bing.com",
            "Content-Type": "application/json",
            "Accept": "application/json",
            "User-Agent": "Math/1 CFNetwork/1121.2.2 Darwin/19.3.0",
        }
        base64 = Image_data.decode('utf-8')
        data = {
            "data": base64,
            "inputForm": "Image",
            "clientInfo": {
                "app": "Math",
                "platform": "ios",
                "configuration": "UnKnown",
                "version": "1.8.0",
                "mkt": "zh-cn"
            },
            "timestamp": int(time.time())
        }
        Html = requests.post(url, headers=headers, json=data)
        res = json.loads(Html.text)
        print(res)
        cnt = 0
        while res["latex"] == '':
            # print("cnt:", cnt)
            Html = requests.post(url, headers=headers, json=data)
            res = json.loads(Html.text)
            cnt += 1
            if cnt == 5:
                break
        if res["latex"] != '':
            return res["latex"]
        else:
            return "Can't recognition"
        
    